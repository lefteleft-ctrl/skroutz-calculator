from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Query
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import io
import re
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import openpyxl

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# --- Models ---

class CalculateRequest(BaseModel):
    uid: str
    wholesale_price: float
    vat_pct: float = 24.0
    profit: float = 0.90
    mgmt_cost: float = 0.0
    coins_quantity: int = 0
    ads_enabled: bool = False

class CalculateResponse(BaseModel):
    model_config = ConfigDict(extra="allow")
    product_name: str
    uid: str
    ean: Optional[str] = None
    category: Optional[str] = None
    wholesale_price: float
    vat_pct: float
    profit: float
    marketplace_commission_pct: float
    advertising_commission_pct: float = 0.0
    ads_enabled: bool = False
    coins_quantity: int = 0
    coins_eur: float = 0.0
    fbs_fee: float
    mgmt_cost: float
    packaging_cost: float
    fbs_final_price: float
    fbs_breakdown: dict
    marketplace_final_price: float
    marketplace_breakdown: dict

class ProductSummary(BaseModel):
    uid: str
    name: str
    category: Optional[str] = None
    manufacturer: Optional[str] = None
    ean: Optional[str] = None
    current_price: Optional[float] = None
    marketplace_commission_pct: Optional[float] = None
    fbs_fee: Optional[float] = None
    weight_kg: Optional[float] = None

class UploadStatus(BaseModel):
    report_listed_count: int = 0
    fbs_products_count: int = 0
    total_products: int = 0
    wholesale_count: int = 0

# --- Helpers ---

def parse_price(val):
    """Parse price values from Excel that may be string like '10,25 €' or float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.replace('€', '').replace(' ', '').replace(',', '.').strip()
        if cleaned and cleaned != '-':
            try:
                return float(cleaned)
            except ValueError:
                return None
    return None

def safe_str(val):
    if val is None:
        return None
    return str(val).strip() if str(val).strip() != '-' else None

def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.replace(',', '.').strip()
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


COIN_COST = 0.0015  # € per coin

def calculate_price(cost, profit, fixed_fee, mp_pct_decimal, vat_pct_decimal, ads_pct_decimal=0.0):
    """
    Correct Skroutz pricing formula.
    Formula: final = (cost + profit + fixed_fee) / (1 - mp% - ads% - (1 - 1/(1+vat%)))
    """
    vat_impact = 1 - 1 / (1 + vat_pct_decimal)
    denominator = 1 - mp_pct_decimal - ads_pct_decimal - vat_impact
    if denominator <= 0:
        return None
    final_price = (cost + profit + fixed_fee) / denominator
    return round(final_price, 2)


def build_breakdown(final_price, cost, profit, fixed_fee, mp_pct_decimal, vat_pct_decimal, fee_label, ads_pct_decimal=0.0):
    """Build a detailed cost breakdown."""
    commission_amount = round(final_price * mp_pct_decimal, 2)
    ads_amount = round(final_price * ads_pct_decimal, 2)
    vat_amount = round(final_price * (1 - 1 / (1 + vat_pct_decimal)), 2)
    net_after_all = round(final_price - commission_amount - ads_amount - vat_amount - fixed_fee, 2)
    real_profit = round(net_after_all - cost, 2)
    return {
        "final_price": final_price,
        "wholesale_price": cost,
        "profit_target": profit,
        fee_label: fixed_fee,
        "commission_pct": round(mp_pct_decimal * 100, 2),
        "commission_amount": commission_amount,
        "ads_pct": round(ads_pct_decimal * 100, 2),
        "ads_amount": ads_amount,
        "vat_pct": round(vat_pct_decimal * 100, 2),
        "vat_amount": vat_amount,
        "net_to_store": net_after_all,
        "real_profit": real_profit,
    }


# --- Excel Upload Endpoints ---

@api_router.post("/upload/report-listed")
async def upload_report_listed(file: UploadFile = File(...)):
    """Upload and parse the report_listed Excel (Marketplace commissions %)."""
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        
        products_updated = 0
        for row_num in range(2, ws.max_row + 1):
            row = {headers[i]: ws.cell(row=row_num, column=i + 1).value for i in range(len(headers)) if headers[i]}
            
            uid = safe_str(row.get('UID'))
            if not uid:
                continue

            doc = {
                "uid": uid,
                "sku_id": safe_str(row.get('SKU ID')),
                "name": safe_str(row.get('Όνομα', '')),
                "current_price": parse_price(row.get('Τιμή')),
                "category": safe_str(row.get('Κατηγορία προϊόντος')),
                "manufacturer": safe_str(row.get('Κατασκευαστής')),
                "ean": safe_str(row.get('ΕΑΝ')),
                "mpn": safe_str(row.get('MPN')),
                "availability": safe_str(row.get('Διαθεσιμότητα')),
                "url": safe_str(row.get('URL προϊόντος')),
                "quantity": safe_float(row.get('Ποσότητα')),
                "marketplace_commission_pct": safe_float(row.get('Προμήθεια Marketplace %')),
                "cps_commission_pct": safe_float(row.get('Προμήθεια CPS %')),
                "advertising_commission_pct": safe_float(row.get('Προμήθεια Διαφήμισης %')),
                "source_report_listed": True,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            # Remove None values
            doc = {k: v for k, v in doc.items() if v is not None}

            await db.products.update_one(
                {"uid": uid},
                {"$set": doc},
                upsert=True
            )
            products_updated += 1

        return {"message": f"Επιτυχής φόρτωση {products_updated} προϊόντων από report_listed", "count": products_updated}
    except Exception as e:
        logging.error(f"Error parsing report_listed: {e}")
        raise HTTPException(status_code=400, detail=f"Σφάλμα ανάγνωσης αρχείου: {str(e)}")


@api_router.post("/upload/fbs-products")
async def upload_fbs_products(file: UploadFile = File(...)):
    """Upload and parse the FBS products active Excel."""
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active

        # Headers are in row 2 for this file
        headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        # Clean header names (remove newlines)
        headers = [h.replace('\n', ' ').strip() if h else None for h in headers]

        products_updated = 0
        for row_num in range(3, ws.max_row + 1):
            row = {}
            for i, h in enumerate(headers):
                if h:
                    row[h] = ws.cell(row=row_num, column=i + 1).value
            
            uid = safe_str(row.get('UID'))
            if not uid:
                continue

            doc = {
                "uid": uid,
                "fbs_name": safe_str(row.get('Όνομα')),
                "fbs_category": safe_str(row.get('Κατηγορία')),
                "fbs_manufacturer": safe_str(row.get('Κατασκευαστής')),
                "fbs_ean": safe_str(row.get('EAN')),
                "fbs_current_price": safe_float(row.get('Τρέχουσα τιμή')),
                "weight_kg": safe_float(row.get('Βάρος (kg)')),
                "fbs_fee": safe_float(row.get('Προμήθεια FBS')),
                "fbs_marketplace_commission_eur": safe_float(row.get('Προμήθεια Marketplace')),
                "management_cost": safe_float(row.get('Κόστος διαχείρισης')),
                "net_to_store": safe_float(row.get('Αποδίδεται στο κατάστημα')),
                "fbs_stock": safe_float(row.get('Απόθεμα FBS')),
                "marketplace_stock": safe_float(row.get('Απόθεμα Marketplace')),
                "sales_last_30": safe_float(row.get('Πωλήσεις marketplace τελευταίες 30 ημέρες')),
                "fbs_sales_last_30": safe_float(row.get('Πωλήσεις FBS τελευταίες 30 ημέρες')),
                "source_fbs_products": True,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            doc = {k: v for k, v in doc.items() if v is not None}

            # Also set name/ean/category from FBS if not already set from report_listed
            set_on_insert = {}
            if doc.get('fbs_name'):
                set_on_insert['name'] = doc['fbs_name']
            if doc.get('fbs_ean'):
                set_on_insert['ean'] = doc['fbs_ean']
            if doc.get('fbs_category'):
                set_on_insert['category'] = doc['fbs_category']
            if doc.get('fbs_manufacturer'):
                set_on_insert['manufacturer'] = doc['fbs_manufacturer']

            update_ops = {"$set": doc}
            if set_on_insert:
                update_ops["$setOnInsert"] = set_on_insert

            await db.products.update_one(
                {"uid": uid},
                update_ops,
                upsert=True
            )
            products_updated += 1

        return {"message": f"Επιτυχής φόρτωση {products_updated} προϊόντων από FBS", "count": products_updated}
    except Exception as e:
        logging.error(f"Error parsing fbs_products: {e}")
        raise HTTPException(status_code=400, detail=f"Σφάλμα ανάγνωσης αρχείου: {str(e)}")


@api_router.post("/upload/wholesale")
async def upload_wholesale_prices(file: UploadFile = File(...)):
    """Upload wholesale pricing Excel and bulk-update products by EAN."""
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))

        # Build EAN -> wholesale_price map from all sheets
        ean_price_map = {}
        for sname in wb.sheetnames:
            ws = wb[sname]
            for r in range(1, ws.max_row + 1):
                a = ws.cell(row=r, column=1).value  # Product name
                b = ws.cell(row=r, column=2).value  # EAN
                c = ws.cell(row=r, column=3).value  # Wholesale price

                # Validate: A is a product name string, B looks like EAN, C is a price
                if not isinstance(a, str) or len(a) < 5:
                    continue
                if b is None:
                    continue
                # Normalize EAN to string (handle float like 5200309854075.0)
                ean = str(b).strip()
                if '.' in ean:
                    ean = ean.split('.')[0]
                if not ean or len(ean) < 7 or not ean.isdigit():
                    continue
                # Parse price
                price = safe_float(c)
                if price is None or price <= 0:
                    continue
                ean_price_map[ean] = price

        if not ean_price_map:
            raise HTTPException(status_code=400, detail="Δεν βρέθηκαν έγκυρα δεδομένα τιμών στο αρχείο")

        # Match by EAN and update wholesale price in DB
        matched = 0
        for ean, price in ean_price_map.items():
            result = await db.products.update_many(
                {"ean": ean},
                {"$set": {
                    "user_wholesale_price": price,
                    "wholesale_source": "excel_upload",
                    "wholesale_updated_at": datetime.now(timezone.utc).isoformat(),
                }}
            )
            if result.modified_count > 0:
                matched += result.modified_count

        return {
            "message": f"Ενημερώθηκαν {matched} προϊόντα από {len(ean_price_map)} τιμές χονδρικής",
            "matched_count": matched,
            "total_prices_in_file": len(ean_price_map),
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error parsing wholesale prices: {e}")
        raise HTTPException(status_code=400, detail=f"Σφάλμα ανάγνωσης αρχείου: {str(e)}")


@api_router.get("/upload-status", response_model=UploadStatus)
async def get_upload_status():
    total = await db.products.count_documents({})
    report_listed = await db.products.count_documents({"source_report_listed": True})
    fbs_products = await db.products.count_documents({"source_fbs_products": True})
    wholesale_count = await db.products.count_documents({"wholesale_source": "excel_upload"})
    return UploadStatus(
        report_listed_count=report_listed,
        fbs_products_count=fbs_products,
        total_products=total,
        wholesale_count=wholesale_count,
    )


# --- Averages Endpoint (for Quick Calculator) ---

@api_router.get("/averages")
async def get_averages():
    """Get average commission/fee values from uploaded data."""
    pipeline_mp = [
        {"$match": {"marketplace_commission_pct": {"$exists": True, "$ne": None}}},
        {"$group": {"_id": None, "avg": {"$avg": "$marketplace_commission_pct"}, "count": {"$sum": 1}}}
    ]
    pipeline_fbs = [
        {"$match": {"fbs_fee": {"$exists": True, "$ne": None}}},
        {"$group": {"_id": None, "avg": {"$avg": "$fbs_fee"}, "count": {"$sum": 1}}}
    ]
    pipeline_mgmt = [
        {"$match": {"management_cost": {"$exists": True, "$ne": None}}},
        {"$group": {"_id": None, "avg": {"$avg": "$management_cost"}, "count": {"$sum": 1}}}
    ]

    mp_result = await db.products.aggregate(pipeline_mp).to_list(1)
    fbs_result = await db.products.aggregate(pipeline_fbs).to_list(1)
    mgmt_result = await db.products.aggregate(pipeline_mgmt).to_list(1)

    return {
        "avg_marketplace_commission_pct": round(mp_result[0]["avg"], 2) if mp_result else 9.24,
        "avg_fbs_fee": round(fbs_result[0]["avg"], 2) if fbs_result else 0.56,
        "avg_management_cost": round(mgmt_result[0]["avg"], 2) if mgmt_result else 0.28,
        "packaging_cost": 0.12,
        "products_count": mp_result[0]["count"] if mp_result else 0,
    }


# --- Quick Calculate Endpoint ---

class QuickCalculateRequest(BaseModel):
    wholesale_price: float
    vat_pct: float = 24.0
    profit: float = 0.90
    mp_pct: Optional[float] = None  # Override average if provided
    fbs_fee: Optional[float] = None
    mgmt_cost: Optional[float] = None

@api_router.post("/quick-calculate")
async def quick_calculate(req: QuickCalculateRequest):
    """Quick calculation using averages from uploaded data."""
    # Get averages if not overridden
    if req.mp_pct is None or req.fbs_fee is None:
        avg_data = await get_averages()
        mp_pct = req.mp_pct if req.mp_pct is not None else avg_data["avg_marketplace_commission_pct"]
        fbs_fee = req.fbs_fee if req.fbs_fee is not None else avg_data["avg_fbs_fee"]
        mgmt_cost = req.mgmt_cost if req.mgmt_cost is not None else avg_data["avg_management_cost"]
    else:
        mp_pct = req.mp_pct
        fbs_fee = req.fbs_fee
        mgmt_cost = req.mgmt_cost if req.mgmt_cost is not None else 0.0

    mp_decimal = mp_pct / 100.0
    vat_decimal = req.vat_pct / 100.0
    packaging = 0.12

    # FBS
    fbs_fixed = fbs_fee + packaging
    fbs_final = calculate_price(req.wholesale_price, req.profit, fbs_fixed, mp_decimal, vat_decimal)
    if fbs_final is None:
        raise HTTPException(status_code=400, detail="Αδύνατος υπολογισμός")
    fbs_breakdown = build_breakdown(fbs_final, req.wholesale_price, req.profit, fbs_fixed, mp_decimal, vat_decimal, "fbs_fee_plus_packaging")

    # Marketplace
    mp_fixed = mgmt_cost
    mp_final = calculate_price(req.wholesale_price, req.profit, mp_fixed, mp_decimal, vat_decimal)
    if mp_final is None:
        raise HTTPException(status_code=400, detail="Αδύνατος υπολογισμός")
    mp_breakdown = build_breakdown(mp_final, req.wholesale_price, req.profit, mp_fixed, mp_decimal, vat_decimal, "management_cost")

    return {
        "wholesale_price": req.wholesale_price,
        "vat_pct": req.vat_pct,
        "profit": req.profit,
        "avg_mp_pct": mp_pct,
        "avg_fbs_fee": fbs_fee,
        "avg_mgmt_cost": mgmt_cost,
        "packaging_cost": packaging,
        "fbs_final_price": fbs_final,
        "fbs_breakdown": fbs_breakdown,
        "marketplace_final_price": mp_final,
        "marketplace_breakdown": mp_breakdown,
    }


# --- Search Endpoint ---

@api_router.get("/products/search")
async def search_products(q: str = Query(..., min_length=1)):
    """Search products by name or EAN."""
    # Try exact EAN match first
    results = await db.products.find(
        {"ean": q},
        {"_id": 0}
    ).to_list(20)

    if not results:
        # Text search on name (case-insensitive regex)
        escaped_q = re.escape(q)
        results = await db.products.find(
            {"$or": [
                {"name": {"$regex": escaped_q, "$options": "i"}},
                {"fbs_name": {"$regex": escaped_q, "$options": "i"}},
                {"ean": {"$regex": escaped_q, "$options": "i"}},
                {"uid": {"$regex": escaped_q, "$options": "i"}},
            ]},
            {"_id": 0}
        ).to_list(20)

    return results


# --- All Products (for table view) - MUST be before /products/{uid} ---

@api_router.get("/products/all")
async def get_all_products():
    """Get all products sorted alphabetically by name."""
    products = await db.products.find(
        {"name": {"$exists": True, "$ne": None}},
        {
            "_id": 0,
            "uid": 1,
            "name": 1,
            "fbs_name": 1,
            "ean": 1,
            "category": 1,
            "fbs_category": 1,
            "manufacturer": 1,
            "fbs_manufacturer": 1,
            "marketplace_commission_pct": 1,
            "advertising_commission_pct": 1,
            "fbs_fee": 1,
            "management_cost": 1,
            "weight_kg": 1,
            "current_price": 1,
            "fbs_current_price": 1,
            "user_wholesale_price": 1,
            "user_coins_quantity": 1,
            "user_ads_enabled": 1,
            "user_profit": 1,
            "user_vat_pct": 1,
        }
    ).sort("name", 1).to_list(5000)
    for p in products:
        if not p.get("name") and p.get("fbs_name"):
            p["name"] = p["fbs_name"]
        if not p.get("category") and p.get("fbs_category"):
            p["category"] = p["fbs_category"]
    return products


@api_router.get("/products/{uid}")
async def get_product(uid: str):
    """Get a single product by UID."""
    product = await db.products.find_one({"uid": uid}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Το προϊόν δεν βρέθηκε")
    return product


# --- Calculate Endpoint ---

@api_router.post("/calculate", response_model=CalculateResponse)
async def calculate_price_endpoint(req: CalculateRequest):
    """Calculate FBS and Marketplace final prices for a product."""
    product = await db.products.find_one({"uid": req.uid}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Το προϊόν δεν βρέθηκε")

    mp_pct = product.get("marketplace_commission_pct")
    if mp_pct is None:
        raise HTTPException(status_code=400, detail="Δεν βρέθηκε ποσοστό προμήθειας Marketplace. Ανεβάστε το report_listed Excel.")

    fbs_fee = product.get("fbs_fee", 0.0) or 0.0
    packaging_cost = 0.12
    ads_pct = product.get("advertising_commission_pct", 0.0) or 0.0

    mp_decimal = mp_pct / 100.0
    vat_decimal = req.vat_pct / 100.0
    ads_decimal = (ads_pct / 100.0) if req.ads_enabled else 0.0
    coins_eur = req.coins_quantity * COIN_COST

    # FBS calculation
    fbs_fixed = fbs_fee + packaging_cost + coins_eur
    fbs_final = calculate_price(req.wholesale_price, req.profit, fbs_fixed, mp_decimal, vat_decimal, ads_decimal)
    if fbs_final is None:
        raise HTTPException(status_code=400, detail="Αδύνατος υπολογισμός - ελέγξτε τις παραμέτρους")
    fbs_breakdown = build_breakdown(fbs_final, req.wholesale_price, req.profit, fbs_fixed, mp_decimal, vat_decimal, "fbs_fee_plus_packaging", ads_decimal)

    # Marketplace calculation
    mp_fixed = req.mgmt_cost + coins_eur
    mp_final = calculate_price(req.wholesale_price, req.profit, mp_fixed, mp_decimal, vat_decimal, ads_decimal)
    if mp_final is None:
        raise HTTPException(status_code=400, detail="Αδύνατος υπολογισμός - ελέγξτε τις παραμέτρους")
    mp_breakdown = build_breakdown(mp_final, req.wholesale_price, req.profit, mp_fixed, mp_decimal, vat_decimal, "management_cost", ads_decimal)

    return CalculateResponse(
        product_name=product.get("name") or product.get("fbs_name", ""),
        uid=req.uid,
        ean=product.get("ean"),
        category=product.get("category") or product.get("fbs_category"),
        wholesale_price=req.wholesale_price,
        vat_pct=req.vat_pct,
        profit=req.profit,
        marketplace_commission_pct=mp_pct,
        advertising_commission_pct=ads_pct,
        ads_enabled=req.ads_enabled,
        coins_quantity=req.coins_quantity,
        coins_eur=round(coins_eur, 4),
        fbs_fee=fbs_fee,
        mgmt_cost=req.mgmt_cost,
        packaging_cost=packaging_cost,
        fbs_final_price=fbs_final,
        fbs_breakdown=fbs_breakdown,
        marketplace_final_price=mp_final,
        marketplace_breakdown=mp_breakdown,
    )


# --- Health ---

@api_router.get("/")
async def root():
    return {"message": "Skroutz Price Calculator API"}


# --- Reverse Calculate ---

class SaveProductSettingsRequest(BaseModel):
    uid: str
    wholesale_price: float
    coins_quantity: int = 0
    ads_enabled: bool = False
    profit: float = 0.90
    vat_pct: float = 24.0

@api_router.post("/save-product-settings")
async def save_product_settings(req: SaveProductSettingsRequest):
    """Save user pricing settings for a product."""
    result = await db.products.update_one(
        {"uid": req.uid},
        {"$set": {
            "user_wholesale_price": req.wholesale_price,
            "user_coins_quantity": req.coins_quantity,
            "user_ads_enabled": req.ads_enabled,
            "user_profit": req.profit,
            "user_vat_pct": req.vat_pct,
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Το προϊόν δεν βρέθηκε")
    return {"message": "Αποθηκεύτηκε", "uid": req.uid}


class ReverseCalculateRequest(BaseModel):
    final_price: float
    wholesale_price: float
    mp_pct: float  # as percentage e.g. 9.24
    fbs_fee: float = 0.56
    vat_pct: float = 24.0
    packaging_cost: float = 0.12

@api_router.post("/reverse-calculate")
async def reverse_calculate(req: ReverseCalculateRequest):
    """Given a final selling price, calculate the real profit."""
    mp_decimal = req.mp_pct / 100.0
    vat_decimal = req.vat_pct / 100.0
    
    commission = req.final_price * mp_decimal
    vat_amount = req.final_price * (1 - 1 / (1 + vat_decimal))
    fbs_fixed = req.fbs_fee + req.packaging_cost
    
    net_to_store = req.final_price - commission - vat_amount - fbs_fixed
    real_profit = net_to_store - req.wholesale_price
    
    return {
        "final_price": req.final_price,
        "wholesale_price": req.wholesale_price,
        "commission_amount": round(commission, 2),
        "commission_pct": req.mp_pct,
        "vat_amount": round(vat_amount, 2),
        "fbs_fee_plus_packaging": round(fbs_fixed, 2),
        "net_to_store": round(net_to_store, 2),
        "real_profit": round(real_profit, 2),
        "is_profitable": real_profit > 0,
    }


# --- Bulk Calculate + Excel Export ---

class BulkProductInput(BaseModel):
    uid: str
    wholesale_price: float

class BulkCalculateRequest(BaseModel):
    products: List[BulkProductInput]
    vat_pct: float = 24.0
    profit: float = 0.90
    mgmt_cost: float = 0.0

@api_router.post("/bulk-calculate")
async def bulk_calculate(req: BulkCalculateRequest):
    """Calculate prices for multiple products at once."""
    uids = [p.uid for p in req.products]
    wholesale_map = {p.uid: p.wholesale_price for p in req.products}
    
    db_products = await db.products.find(
        {"uid": {"$in": uids}},
        {"_id": 0}
    ).to_list(5000)
    
    db_map = {p["uid"]: p for p in db_products}
    results = []
    
    vat_decimal = req.vat_pct / 100.0
    packaging = 0.12
    
    for uid, cost in wholesale_map.items():
        product = db_map.get(uid)
        if not product:
            continue
        mp_pct = product.get("marketplace_commission_pct")
        if mp_pct is None:
            continue
        
        mp_decimal = mp_pct / 100.0
        fbs_fee = product.get("fbs_fee", 0.0) or 0.0
        
        # FBS
        fbs_fixed = fbs_fee + packaging
        fbs_final = calculate_price(cost, req.profit, fbs_fixed, mp_decimal, vat_decimal)
        
        # Marketplace
        mp_fixed = req.mgmt_cost
        mp_final = calculate_price(cost, req.profit, mp_fixed, mp_decimal, vat_decimal)
        
        if fbs_final and mp_final:
            fbs_bd = build_breakdown(fbs_final, cost, req.profit, fbs_fixed, mp_decimal, vat_decimal, "fbs_fee_plus_packaging")
            mp_bd = build_breakdown(mp_final, cost, req.profit, mp_fixed, mp_decimal, vat_decimal, "management_cost")
            
            results.append({
                "uid": uid,
                "name": product.get("name") or product.get("fbs_name", ""),
                "ean": product.get("ean"),
                "category": product.get("category") or product.get("fbs_category"),
                "wholesale_price": cost,
                "marketplace_commission_pct": mp_pct,
                "fbs_fee": fbs_fee,
                "fbs_final_price": fbs_final,
                "fbs_real_profit": fbs_bd["real_profit"],
                "marketplace_final_price": mp_final,
                "mp_real_profit": mp_bd["real_profit"],
            })
    
    return {"results": results, "count": len(results)}


from fastapi.responses import StreamingResponse

@api_router.post("/export-excel")
async def export_excel(req: BulkCalculateRequest):
    """Generate and download Excel with all calculations."""
    # First do bulk calculation
    uids = [p.uid for p in req.products]
    wholesale_map = {p.uid: p.wholesale_price for p in req.products}
    
    db_products = await db.products.find(
        {"uid": {"$in": uids}},
        {"_id": 0}
    ).to_list(5000)
    
    db_map = {p["uid"]: p for p in db_products}
    
    vat_decimal = req.vat_pct / 100.0
    packaging = 0.12
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Κοστολόγηση Skroutz"
    
    # Headers
    headers = [
        "Όνομα", "EAN", "UID", "Κατηγορία",
        "Χονδρική (€)", "Κέρδος (€)", "ΦΠΑ %",
        "Προμήθεια MP %", "FBS Fee (€)", "Συσκευασία (€)",
        "FBS Τελική Τιμή (€)", "FBS Προμήθεια (€)", "FBS ΦΠΑ (€)", "FBS Καθαρό (€)", "FBS Κέρδος (€)",
        "MP Τελική Τιμή (€)", "MP Προμήθεια (€)", "MP ΦΠΑ (€)", "MP Καθαρό (€)", "MP Κέρδος (€)",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    
    row_num = 2
    for uid, cost in wholesale_map.items():
        product = db_map.get(uid)
        if not product:
            continue
        mp_pct = product.get("marketplace_commission_pct")
        if mp_pct is None:
            continue
        
        mp_decimal = mp_pct / 100.0
        fbs_fee = product.get("fbs_fee", 0.0) or 0.0
        fbs_fixed = fbs_fee + packaging
        mp_fixed = req.mgmt_cost
        
        fbs_final = calculate_price(cost, req.profit, fbs_fixed, mp_decimal, vat_decimal)
        mp_final = calculate_price(cost, req.profit, mp_fixed, mp_decimal, vat_decimal)
        
        if not fbs_final or not mp_final:
            continue
        
        fbs_bd = build_breakdown(fbs_final, cost, req.profit, fbs_fixed, mp_decimal, vat_decimal, "fbs_fee_plus_packaging")
        mp_bd = build_breakdown(mp_final, cost, req.profit, mp_fixed, mp_decimal, vat_decimal, "management_cost")
        
        name = product.get("name") or product.get("fbs_name", "")
        row_data = [
            name, product.get("ean"), uid, product.get("category") or product.get("fbs_category"),
            cost, req.profit, req.vat_pct,
            mp_pct, fbs_fee, packaging,
            fbs_final, fbs_bd["commission_amount"], fbs_bd["vat_amount"], fbs_bd["net_to_store"], fbs_bd["real_profit"],
            mp_final, mp_bd["commission_amount"], mp_bd["vat_amount"], mp_bd["net_to_store"], mp_bd["real_profit"],
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=val)
        row_num += 1
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=skroutz_pricing.xlsx"}
    )


# --- Orders / Profit Calculation ---

COIN_COST_EUR = 0.0015

@api_router.post("/upload/orders")
async def upload_orders(file: UploadFile = File(...)):
    """Parse orders Excel (.xls), match by EAN, calculate profit per product."""
    import xlrd

    contents = await file.read()
    try:
        wb = xlrd.open_workbook(file_contents=contents)
    except Exception:
        raise HTTPException(status_code=400, detail="Μη έγκυρο αρχείο .xls")

    ws = wb.sheet_by_index(0)

    # Parse orders: extract EAN, quantity, skroutz selling price
    order_items = []
    current_order = {}
    for r in range(1, ws.nrows):
        date_val = ws.cell_value(r, 0)
        product_name = ws.cell_value(r, 20)

        if date_val:
            # Order header row
            status = str(ws.cell_value(r, 4)).strip()
            current_order = {"status": status}
        elif product_name:
            # Product row
            ean = str(ws.cell_value(r, 24)).strip()
            if '.' in ean:
                ean = ean.split('.')[0]

            qty_val = ws.cell_value(r, 22)
            qty = int(float(qty_val)) if qty_val else 1

            price_str = str(ws.cell_value(r, 21)).replace(',', '.')
            try:
                skroutz_price = float(price_str)
            except (ValueError, TypeError):
                skroutz_price = 0

            status = current_order.get("status", "")
            if "Ακυρωμένη" in status or "Επιστροφή" in status:
                continue

            order_items.append({
                "name_from_excel": str(product_name).strip(),
                "ean": ean,
                "quantity": qty,
                "skroutz_sell_price": skroutz_price,
            })

    if not order_items:
        raise HTTPException(status_code=400, detail="Δεν βρέθηκαν παραγγελίες στο αρχείο")

    # Get unique EANs and fetch from DB
    ean_set = set(item["ean"] for item in order_items if item["ean"])
    products_cursor = db.products.find(
        {"ean": {"$in": list(ean_set)}},
        {"_id": 0}
    )
    ean_to_product = {}
    async for p in products_cursor:
        ean_to_product[p.get("ean")] = p

    # Calculate profit for each order item
    results = []
    total_profit = 0
    total_revenue = 0
    matched_count = 0
    unmatched = []

    for item in order_items:
        product = ean_to_product.get(item["ean"])
        if not product:
            unmatched.append(item["name_from_excel"])
            continue

        matched_count += 1
        my_price = product.get("current_price") or product.get("fbs_current_price") or 0
        wholesale = product.get("user_wholesale_price") or 0
        mp_pct = product.get("marketplace_commission_pct") or 0
        fbs_fee = product.get("fbs_fee") or 0
        ad_pct = (product.get("advertising_commission_pct") or 0) if product.get("user_ads_enabled") else 0
        coins_qty = product.get("user_coins_quantity") or 0
        vat_pct = product.get("user_vat_pct") or 24.0

        vat_decimal = vat_pct / 100.0
        commission = my_price * (mp_pct / 100.0)
        ad_cost = my_price * (ad_pct / 100.0)
        vat_amount = my_price * (1 - 1 / (1 + vat_decimal))
        coins_cost = coins_qty * COIN_COST_EUR
        fixed_costs = fbs_fee + 0.12 + coins_cost

        profit_per_unit = my_price - commission - ad_cost - vat_amount - fixed_costs - wholesale
        profit_per_unit = round(profit_per_unit, 2)
        total_for_item = round(profit_per_unit * item["quantity"], 2)

        price_mismatch = abs(my_price - item["skroutz_sell_price"]) > 0.02 if my_price > 0 and item["skroutz_sell_price"] > 0 else False

        total_profit += total_for_item
        total_revenue += round(my_price * item["quantity"], 2)

        results.append({
            "name": product.get("name", item["name_from_excel"]),
            "ean": item["ean"],
            "quantity": item["quantity"],
            "my_price": round(my_price, 2),
            "skroutz_price": round(item["skroutz_sell_price"], 2),
            "wholesale": round(wholesale, 2),
            "commission": round(commission, 2),
            "vat_amount": round(vat_amount, 2),
            "ad_cost": round(ad_cost, 2),
            "fbs_fee": round(fbs_fee + 0.12, 2),
            "coins_cost": round(coins_cost, 4),
            "profit_per_unit": profit_per_unit,
            "total_profit": total_for_item,
            "price_mismatch": price_mismatch,
        })

    return {
        "results": results,
        "summary": {
            "total_items": len(results),
            "total_quantity": sum(r["quantity"] for r in results),
            "total_revenue": round(total_revenue, 2),
            "total_profit": round(total_profit, 2),
            "matched": matched_count,
            "unmatched_count": len(unmatched),
            "unmatched_names": unmatched[:20],
        }
    }


@api_router.post("/calculate-manual-profit")
async def calculate_manual_profit(items: List[dict]):
    """Calculate profit for manually entered barcode+quantity pairs."""
    eans = [str(item.get("ean", "")).strip() for item in items if item.get("ean")]
    products_cursor = db.products.find({"ean": {"$in": eans}}, {"_id": 0})
    ean_to_product = {}
    async for p in products_cursor:
        ean_to_product[p.get("ean")] = p

    results = []
    total_profit = 0
    for item in items:
        ean = str(item.get("ean", "")).strip()
        qty = int(item.get("quantity", 1))
        product = ean_to_product.get(ean)
        if not product:
            results.append({"ean": ean, "quantity": qty, "matched": False, "name": "Δεν βρέθηκε"})
            continue

        my_price = product.get("current_price") or product.get("fbs_current_price") or 0
        wholesale = product.get("user_wholesale_price") or 0
        mp_pct = product.get("marketplace_commission_pct") or 0
        fbs_fee = product.get("fbs_fee") or 0
        ad_pct = (product.get("advertising_commission_pct") or 0) if product.get("user_ads_enabled") else 0
        coins_qty = product.get("user_coins_quantity") or 0
        vat_pct = product.get("user_vat_pct") or 24.0

        vat_decimal = vat_pct / 100.0
        commission = my_price * (mp_pct / 100.0)
        ad_cost = my_price * (ad_pct / 100.0)
        vat_amount = my_price * (1 - 1 / (1 + vat_decimal))
        coins_cost = coins_qty * COIN_COST_EUR
        fixed_costs = fbs_fee + 0.12 + coins_cost

        profit_per_unit = round(my_price - commission - ad_cost - vat_amount - fixed_costs - wholesale, 2)
        total_for_item = round(profit_per_unit * qty, 2)
        total_profit += total_for_item

        results.append({
            "name": product.get("name", ""),
            "ean": ean,
            "quantity": qty,
            "matched": True,
            "my_price": round(my_price, 2),
            "wholesale": round(wholesale, 2),
            "profit_per_unit": profit_per_unit,
            "total_profit": total_for_item,
        })

    return {"results": results, "total_profit": round(total_profit, 2)}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
